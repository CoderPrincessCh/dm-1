import requests
import re
import csv
from time import sleep
from html import unescape
import matplotlib.pyplot as plt
import pandas as pd
import matplotlib
import certifi
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import uuid
from datetime import datetime

DRAMALIST_API = "https://www.missevan.com/dramaapi/filter"
# DRAMALIST_API = "https://app.missevan.com/drama/filter"
SEARCH_API = "https://www.missevan.com/dramaapi/search"
# app_headers = {
#     "User-Agent": "MissEvanApp/6.4.0 (Android;9;Asus ASUS_I003DD ASUS_I003DD)",
#     "channel": "missevan",
#     "Accept": "application/json",
#     "Cookie": "equip_id=680089c7-c0e7-af6a-acab-1d41fcc32ea0; buvid=XZ223262BF341FCFF274134D12230B4746235; device_token=v1|5WodYAe7rq9IOf2uQRqVC91Rt5tUEKq3IWEle1lIumjiWuudxLWrmn4znl182dEeP",
#     "Authorization": "MissEvan pwRVqLHmZB0T2CmfYNwRmbVBz5iqA2Op2xm2cdgFzfs=",
#     "X-M-Date": datetime.utcnow().isoformat(timespec='milliseconds') + "Z",
#     "X-M-Nonce": str(uuid.uuid4()),
#     "bili-bridge-engine": "cronet",
#     "Accept-Encoding": "gzip, deflate, br"
# }
headers = {
    "User-Agent": "Mozilla/5.0"
}
requests.get(DRAMALIST_API, headers=headers, verify=certifi.where())

def clean_html(raw_html):
    # 去除 HTML 标签
    text = re.sub(r"<.*?>", "", raw_html)
    return unescape(text)

def extract_producers(abstract_html: str):
    text = clean_html(abstract_html)

    # 提取“联合出品”、“出品”、“联合制作” 等句子
    match = re.search(r"(猫耳FM.*?出品|.*?联合出品|.*?出品)", text)
    producers = match.group(0).strip() if match else ""

    has_maoer = "猫耳FM" in producers
    return producers, "是" if has_maoer else "否"

def extract_all_producers(abstract_html: str):
    text = clean_html(abstract_html)

    # 先找所有包含“出品”、“联合制作”等关键词的短句
    # 这里匹配连续不超过20个非换行字符，且包含“出品”或“制作”相关关键词
    pattern = r"([\u4e00-\u9fa5A-Za-z0-9·、，, ]{1,30}?(联合出品|联合制作|出品|制作|协作)[\u4e00-\u9fa5A-Za-z0-9·、，, ]{0,10})"
    
    matches = re.findall(pattern, text)

    # matches 是 (匹配短句, 关键词) 的元组列表，取第0个元素
    producers_list = [m[0].strip(" ,，") for m in matches]

    # 去重且保持顺序
    seen = set()
    producers = []
    for p in producers_list:
        if p not in seen:
            seen.add(p)
            producers.append(p)

    for p in producers_list:
        # 过滤包含以下词的条目
        if ("原著" in p) or ("广播剧" in p) or ("作者" in p):
            continue
        if p not in seen:
            seen.add(p)
            producers.append(p)

    # 是否含“猫耳FM”
    has_maoer = any("猫耳FM" in p for p in producers)

    # 返回所有出品方列表，和“是否猫耳FM”标识
    return producers, "是" if has_maoer else "否"

def add_category_by_abstract(abstract_html: str):
    """
    根据 abstract 字段给每条数据加 category 字段：
    包含“广播剧”就是广播剧，其他归为“有声剧及其他”
    """
    text = clean_html(abstract_html)
    if "有声剧" in text:
        return "有声剧"
    else:
        return "广播剧"


def chinese_to_arabic(cn: str) -> int | None:
    cn_num = {
        '零': 0, '一': 1, '二': 2, '两': 2, '三': 3, '四': 4,
        '五': 5, '六': 6, '七': 7, '八': 8, '九': 9
    }
    cn_unit = {'十': 10, '百': 100, '千': 1000}

    result = 0
    unit = 1  # 当前单位
    tmp = 0   # 临时数字存储，用于处理个位和单位结合

    for i in reversed(cn):
        if i in cn_unit:
            unit = cn_unit[i]
            if tmp == 0:  # 如“十”单独，代表10
                tmp = 1
            result += tmp * unit
            tmp = 0
            unit = 1
        elif i in cn_num:
            tmp = cn_num[i]
        else:
            return None  # 遇到不认识的字符返回None

    result += tmp * unit
    return result if result != 0 else None


def extract_episode_counts(abstract_html: str):
    text = clean_html(abstract_html)

    main_eps = None
    total_eps = None
    extra_eps = 0

    # 优先匹配：9集正剧 + 4个番外
    combo_match = re.search(
        r"(\d+)\s*[集话期]+\s*正剧\s*[,、，\s]*\s*(\d+)\s*[集话期]*\s*(番外|花絮|小剧场|福利)?",
        text
    )
    if combo_match:
        main_eps = int(combo_match.group(1))
        extra_eps = int(combo_match.group(2))
        total_eps = main_eps + extra_eps

    # 匹配：共X集正剧 / 正剧X集
    if main_eps is None:
        alt_main_match = re.search(r"(?:正剧\s*共?|共\s*)?(\d+)\s*[集话期]\s*正剧", text)
        if alt_main_match:
            main_eps = int(alt_main_match.group(1))
            total_eps = main_eps

    # 匹配：295+集正剧
    if main_eps is None:
        plus_match = re.search(r"(?:正剧\s*共|共)?\s*(\d+)\+\s*[集话期]\s*正剧", text)
        if plus_match:
            main_eps = int(plus_match.group(1))
            total_eps = None  # 不确定

    # 匹配：共X集（含番外）
    if main_eps is None:
        total_match = re.search(r"共\s*(\d+)\s*[集话期]", text)
        if total_match:
            total_eps = int(total_match.group(1))

            extra_match = re.search(r"含\s*(\d+)\s*[集话期]?.*?(番外|花絮|小剧场|福利)", text)
            extra_eps = int(extra_match.group(1)) if extra_match else 0

            main_eps = total_eps - extra_eps

    # 匹配：共X期 / 第一季X集
    if main_eps is None:
        fallback_match = re.search(r"(?:共|第一季)?\s*(\d{1,4})\s*[集期话]", text)
        if fallback_match:
            main_eps = int(fallback_match.group(1))
            total_eps = main_eps

    # 匹配：中文数字 “十六集”
    if main_eps is None:
      zh_match = re.search(r"(?:共)?([零一二三四五六七八九十百千两]+)[集期话]", text)
      if zh_match:
          cn_number = zh_match.group(1)
          main_eps = chinese_to_arabic(cn_number)
          total_eps = main_eps


    # 匹配：“包含正剧14期”
    # 新增：更宽松匹配“包含正剧14期、小剧场...”结构
    if main_eps is None:
        contain_main_match = re.search(r"包含\s*正剧\s*([0-9零一二三四五六七八九十百千两]+)[集话期]", text)
        if contain_main_match:
            raw = contain_main_match.group(1)
            main_eps = int(raw) if raw.isdigit() else chinese_to_arabic(raw)
            total_eps = main_eps  # 先假设总集数等于正剧数


    # 匹配“前X集免费”/“首X集限免”
    free_match = re.search(r"(?:前|首)\s*([0-9零一二三四五六七八九十百千两]+)\s*[集话期]?.*?(免费|限免)", text)
    # free_eps = int(free_match.group(1)) if free_match else 0
    if free_match:
        raw = free_match.group(1)
        if raw.isdigit():
            free_eps = int(raw)
        else:
            free_eps = chinese_to_arabic(raw) or 0
    else:
        free_eps = 0


    # 新增匹配：第一季X期 / 第一季X集，兼容“第一季10期，共229钻石”这种格式
    if main_eps is None:
        season_match = re.search(r"第一季\s*(\d{1,4})\s*[集期话]", text)
        if season_match:
            main_eps = int(season_match.group(1))
            total_eps = main_eps

    # 付费集 = 正剧 - 免费
    if main_eps is not None:
        paid_eps = max(main_eps - free_eps, 0)
    else:
        paid_eps = None

    return total_eps, main_eps, paid_eps

# web请求
def fetch_drama_list_page(page: int):
    # url = f"{DRAMALIST_API}?filters=0_0_0_0&page={page}"
    # url = f"{DRAMALIST_API}?filters=2_0_4_0_0_2_0&page={page}"
    url = f"{DRAMALIST_API}?filters=0_0_0_0&page={page}&order=1&page_size=20&type=3"
    resp = requests.get(url, headers=headers, timeout=10)
    resp.raise_for_status()
    return resp.json()["info"]["Datas"]

# APP请求（失败
# def fetch_drama_list_page(page: int):
#     url = f"{DRAMALIST_API}?filters=2_0_4_0_0_2_0&page={page}"
#     resp = requests.get(url, headers=app_headers, timeout=10, verify=False)
#     resp.raise_for_status()
#     return resp.json()["info"]["Datas"]

def extract_cover_date(cover_url: str) -> str:
    """
    从 cover 链接中提取日期，格式如：202501，返回 "2025-01"
    """
    match = re.search(r"/(\d{6})/", cover_url)
    if match:
        raw = match.group(1)
        year = raw[:4]
        month = raw[4:]
        return f"{year}-{month}"
    return "未知"


def fetch_all_seasons_by_name(drama_name: str):
    """
    根据主剧名搜索，返回所有匹配剧（包含多季），
    返回列表，每个元素是 dict，包含价格、集数、名字、abstract
    """
    params = {"s": drama_name, "page": 1}
    resp = requests.get(SEARCH_API, headers=headers, params=params, timeout=10)
    resp.raise_for_status()
    candidates = resp.json()["info"]["Datas"]

    results = []
    for item in candidates:
        name = item.get("name", "")
        price_diamond = item.get("price", 0)

        # 只保留价格大于0的
        if price_diamond <= 0:
            continue

        # 过滤只要包含主剧名的结果
        if drama_name in name:
            abstract = item.get("abstract", "")
            producers, is_maoer = extract_all_producers(abstract)
            total_eps, main_eps, paid_eps = extract_episode_counts(abstract)
            cover_url = item.get("cover", "")
            cover_date = extract_cover_date(cover_url)
            category = add_category_by_abstract(abstract)
            results.append({
                "剧名": name,
                "价格钻石": price_diamond,
                "总集数": total_eps,
                "总集数（不含番外）": main_eps,
                "abstract": abstract,
                "付费集数": paid_eps,
                "备注": "；".join(producers),
                "是否猫耳FM": is_maoer,
                "封面时间": cover_date,
                "类型": category,
            })

    if not results:
        print(f"⚠️ 未找到匹配剧目: {drama_name}")
    return results


def clean_drama_name(name: str) -> str:
    """
    去掉“第X季”、“上季”、“下季”、“全一季”、“合集”等后缀，仅保留主剧名
    """
    return re.sub(r"[\s·]*(第[一二三四五六七八九十1234567890]+季|全[一二三四五六七八九十1234567890]+季|上季|下季|合集)$", "", name).strip()


def process_one_drama(drama):
    raw_name = drama["name"]
    drama_name = clean_drama_name(raw_name)
    drama_id = drama["id"]
    # cover_url = drama.get("cover", "")
    # cover_date = extract_cover_date(cover_url)

    all_seasons = fetch_all_seasons_by_name(drama_name)
    results = []

    if not all_seasons:
        print(f"⚠️ {raw_name} 未找到任何季，跳过")
        return []
    
    # 获取第一季封面时间
    first_season_time = all_seasons[0].get("封面时间", "未知")

    for season in all_seasons:
        cover_date = season.get("封面时间", "未知")
        # 跳过简介中包含“原创”的剧
        if "原创" in season["abstract"]:
            continue
        price_rmb = season["价格钻石"] / 10
        main_eps = season["总集数（不含番外）"]
        paid_eps = season["付费集数"]
        producers = season.get("备注", [])
        is_maoer = season.get("是否猫耳FM", "否")
        category = season.get("类型")

        # 跳过价格低于 50 钻石（即 5 元）的剧
        if season["价格钻石"] < 50:
            continue

        if paid_eps is None or paid_eps == 0:
            unit_price = 0
        else:
            unit_price = round(price_rmb / paid_eps, 2)

        if main_eps is None or main_eps == 0:
            avg_price = 0
            print(f"⚠️ {season['剧名']} 无法计算集均价（正剧未知或为0）")
        else:
            avg_price = round(price_rmb / main_eps, 2)

        results.append({
            "封面时间": cover_date,
            "剧名": season["剧名"],
            "剧ID": drama_id,
            "总集数（含番外）": season["总集数"] if season["总集数"] else "未知",
            "总集数（不含番外）": main_eps if main_eps else "未知",
            "总价（元）": round(price_rmb, 2),
            "集均价（元）": avg_price,
            "付费集数": paid_eps if paid_eps else "未知",
            "付费集单价（元）": unit_price,
            "备注": producers,
            "是否猫耳FM": is_maoer,
            "first_season_time": first_season_time,  # 新增字段，只用于排序
            "类型": category,
            
        })

    return results

def split_by_abstract_type(results):
    broadcast_drama = []
    other_drama = []

    for item in results:
        abstract = item.get("abstract", "")
        if "广播剧" in abstract:
            broadcast_drama.append(item)
        else:
            # 有声剧或者其他都归到这里
            other_drama.append(item)
    
    return broadcast_drama, other_drama

def apply_style_to_sheet(ws):
    header_font = Font(bold=True)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    year_row_fill = PatternFill("solid", fgColor="D9E1F2")  # 淡蓝色

    # 标题行样式
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_wrap

    max_col = ws.max_column

    row_idx = 2
    while row_idx <= ws.max_row:
        first_cell_value = ws.cell(row=row_idx, column=1).value
        if isinstance(first_cell_value, str) and first_cell_value.startswith("📅 "):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
            merged_cell = ws.cell(row=row_idx, column=1)
            merged_cell.font = header_font
            merged_cell.alignment = center_wrap
            merged_cell.fill = year_row_fill
            row_idx += 1
        else:
            for col_idx in range(1, max_col + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = center_wrap
            row_idx += 1

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2


def export_with_style(filename, df):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    header_font = Font(bold=True)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    year_row_fill = PatternFill("solid", fgColor="D9E1F2")  # 淡蓝色

    # 设置标题行样式（第一行）
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_wrap

    max_col = ws.max_column

    # 从第二行开始设置样式，合并“📅 年份”行
    row_idx = 2
    while row_idx <= ws.max_row:
        first_cell_value = ws.cell(row=row_idx, column=1).value
        if isinstance(first_cell_value, str) and first_cell_value.startswith("📅 "):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
            merged_cell = ws.cell(row=row_idx, column=1)
            merged_cell.font = header_font
            merged_cell.alignment = center_wrap
            merged_cell.fill = year_row_fill
            row_idx += 1
        else:
            # 普通数据行，全部居中
            for col_idx in range(1, max_col + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = center_wrap
            row_idx += 1

    # 自动调整列宽
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(filename)
    print(f"样式应用完成：{filename}")

def main(export_csv=True):
    result = []
    page = 1
    max_pages = 200  # 最大抓取页数限制
    processed_main_titles = set()

    while True:
        # if page > max_pages:
        #     print(f"📢 已达到最大抓取页数 {max_pages}，结束抓取。")
        #     break

        print(f"📦 正在抓取第 {page} 页剧列表...")
        try:
            dramas = fetch_drama_list_page(page)
        except Exception as e:
            print(f"❗ 第 {page} 页获取失败：{e}")
            break

        if not dramas:
            print("✅ 抓取完毕")
            break

        for drama in dramas:
          if drama.get("pay_type") != 2:
              continue
            
          # 获取主剧名（去掉第几季）
          raw_name = drama.get("name", "")
          main_title = clean_drama_name(raw_name)

          if main_title in processed_main_titles:
              continue  # ✅ 已处理过该主剧，跳过
          processed_main_titles.add(main_title)

          season_results = process_one_drama(drama)
          if season_results:
              result.extend(season_results)
              for data in season_results:
                  print(f"✅ {data['封面时间']} | {data['剧名']} | 类型：{data['类型']} | 总集数（含番外）{data['总集数（含番外）']}集 | 总集数（不含番外）{data['总集数（不含番外）']}集 | 付费{data['付费集数']}集 | 总价（元）：¥{data['总价（元）']} | 集均价：¥{data['集均价（元）']} | 付费集单价：¥{data['付费集单价（元）']} | 备注：{data['备注']}")
          sleep(0.3)
        page += 1
        sleep(0.5)

        excel_file = "广播剧_有声剧_分表导出.xlsx"
        if export_csv and result:
            df = pd.DataFrame(result)
    
            # 统一添加年份列并排序
            df["年份"] = df["first_season_time"].str[:4]
            df = df.sort_values(by=["first_season_time", "封面时间"])
    
            export_cols = ["封面时间", "剧名", "总价（元）", "总集数（不含番外）",
                           "集均价（元）", "付费集数", "付费集单价（元）"]
    
            def build_grouped_rows(sub_df):
                grouped_rows = []
                grouped = sub_df.groupby("年份")
                for year, group in grouped:
                    # 插入年份标题行
                    grouped_rows.append({"剧名": f"📅 {year} 年"})
                    grouped_rows.extend(group[export_cols].to_dict(orient="records"))
                return grouped_rows
    
            # 广播剧
            radio_df = df[df["类型"] == "广播剧"]
            radio_rows = build_grouped_rows(radio_df)
            radio_final_df = pd.DataFrame(radio_rows)
    
            # 有声剧及其他
            other_df = df[df["类型"] != "广播剧"]
            other_rows = build_grouped_rows(other_df)
            other_final_df = pd.DataFrame(other_rows)
    
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                radio_final_df.to_excel(writer, sheet_name="广播剧", index=False)
                other_final_df.to_excel(writer, sheet_name="有声剧及其他", index=False)
    
        # 分别应用样式到两个 sheet
        # openpyxl 默认打开的是第一个 sheet，第二个需要切换
        wb = openpyxl.load_workbook(excel_file)

        # 应用广播剧sheet样式
        ws_radio = wb["广播剧"]
        apply_style_to_sheet(ws_radio)

        # 应用有声剧及其他sheet样式
        ws_other = wb["有声剧及其他"]
        apply_style_to_sheet(ws_other)

        wb.save(excel_file)
        print(f"📄 分组导出并美化完成：{excel_file}")

    return result

# 设置中文字体支持（如使用微软雅黑）
matplotlib.rcParams['font.family'] = ['Microsoft YaHei']
matplotlib.rcParams['axes.unicode_minus'] = False  # 负号正常显示

# def draw_price_charts(csv_file="全部付费广播剧_集均价统计.csv"):
# # def draw_price_charts(csv_file="全部付费广播剧_集均价统计.csv", top_n=30):
#     df = pd.read_csv(csv_file)

#     # 过滤掉无法转数字的项
#     df = df[pd.to_numeric(df["集均价（元）"], errors='coerce').notna()]
#     df = df[pd.to_numeric(df["付费集单价（元）"], errors='coerce').notna()]

#     df["集均价（元）"] = df["集均价（元）"].astype(float)
#     df["付费集单价（元）"] = df["付费集单价（元）"].astype(float)

#     # 排序后取前 N 项
#     top_avg = df.sort_values("集均价（元）", ascending=False).head(top_n)
#     top_unit = df.sort_values("付费集单价（元）", ascending=False).head(top_n)

#     # ✅ 图1：集均价
#     plt.figure(figsize=(12, 8))
#     bars = plt.barh(top_avg["剧名"], top_avg["集均价（元）"], color="skyblue")
#     plt.xlabel("集均价（元）")
#     # plt.title(f"集均价 Top {top_n}")
#     plt.title(f"集均价 Top")
#     plt.gca().invert_yaxis()
#     plt.yticks(fontsize=10)
#     plt.tight_layout()
#     plt.savefig("集均价_TOP.png", dpi=200)
#     plt.close()

#     # ✅ 图2：付费集单价
#     plt.figure(figsize=(12, 8))
#     bars = plt.barh(top_unit["剧名"], top_unit["付费集单价（元）"], color="salmon")
#     plt.xlabel("付费集单价（元）")
#     plt.title(f"付费集单价 Top")
#     # plt.title(f"付费集单价 Top {top_n}")
#     plt.gca().invert_yaxis()
#     plt.yticks(fontsize=10)
#     plt.tight_layout()
#     plt.savefig("付费集单价_TOP.png", dpi=200)
#     plt.close()

#     print("📊 图表已生成：集均价_TOP.png，付费集单价_TOP.png")

# def draw_price_charts(csv_file="全部付费广播剧_集均价统计.csv"):
#     df = pd.read_csv(csv_file)

#     # 清洗数据
#     df = df[pd.to_numeric(df["集均价（元）"], errors='coerce').notna()]
#     df = df[pd.to_numeric(df["付费集单价（元）"], errors='coerce').notna()]
#     df["集均价（元）"] = df["集均价（元）"].astype(float)
#     df["付费集单价（元）"] = df["付费集单价（元）"].astype(float)

#     # 不截取前 30，直接全量绘图
#     sorted_avg = df.sort_values("集均价（元）", ascending=False)
#     sorted_unit = df.sort_values("付费集单价（元）", ascending=False)

#     # 图1：全部集均价
#     plt.figure(figsize=(12, max(6, 0.3 * len(sorted_avg))))
#     plt.barh(sorted_avg["剧名"], sorted_avg["集均价（元）"], color="skyblue")
#     plt.xlabel("集均价（元）")
#     plt.title("全部广播剧 - 集均价")
#     plt.gca().invert_yaxis()
#     plt.yticks(fontsize=9)
#     plt.tight_layout()
#     plt.savefig("集均价_ALL.png", dpi=200)
#     plt.close()

#     # 图2：全部付费集单价
#     plt.figure(figsize=(12, max(6, 0.3 * len(sorted_unit))))
#     plt.barh(sorted_unit["剧名"], sorted_unit["付费集单价（元）"], color="salmon")
#     plt.xlabel("付费集单价（元）")
#     plt.title("全部广播剧 - 付费集单价")
#     plt.gca().invert_yaxis()
#     plt.yticks(fontsize=9)
#     plt.tight_layout()
#     plt.savefig("付费集单价_ALL.png", dpi=200)
#     plt.close()

#     print("📊 图表已生成：集均价_ALL.png，付费集单价_ALL.png")

def draw_price_charts_paged(csv_file="全部付费广播剧_集均价统计.csv", page_size=50):
    df = pd.read_csv(csv_file)

    # 清洗数据
    df = df[pd.to_numeric(df["集均价（元）"], errors='coerce').notna()]
    df = df[pd.to_numeric(df["付费集单价（元）"], errors='coerce').notna()]
    df["集均价（元）"] = df["集均价（元）"].astype(float)
    df["付费集单价（元）"] = df["付费集单价（元）"].astype(float)

    # 按集均价排序和分组分页
    sorted_avg = df.sort_values("集均价（元）", ascending=False).reset_index(drop=True)
    sorted_unit = df.sort_values("付费集单价（元）", ascending=False).reset_index(drop=True)

    def plot_paged(data, value_col, title_prefix, color, filename_prefix):
        total = len(data)
        pages = (total + page_size - 1) // page_size

        for i in range(pages):
            start = i * page_size
            end = min((i + 1) * page_size, total)
            chunk = data.iloc[start:end]

            plt.figure(figsize=(12, 0.4 * len(chunk) + 2))
            plt.barh(chunk["剧名"], chunk[value_col], color=color)
            plt.xlabel(value_col)
            plt.title(f"{title_prefix}（第 {i+1} 页，共 {pages} 页）")
            plt.gca().invert_yaxis()
            plt.yticks(fontsize=9)
            plt.tight_layout()
            plt.savefig(f"{filename_prefix}_Page_{i+1}.png", dpi=200)
            plt.close()

    # 分页绘制两类图
    plot_paged(sorted_avg, "集均价（元）", "广播剧集均价 Top", "skyblue", "集均价")
    plot_paged(sorted_unit, "付费集单价（元）", "广播剧付费集单价 Top", "salmon", "付费集单价")

    print("📊 分页图表已生成（如：集均价_Page_1.png，付费集单价_Page_1.png 等）")




if __name__ == "__main__":
    RUN_MAIN = True     # 改为 True 就会运行 main()
    RUN_DRAW = False      # 改为 False 就不绘图

    if RUN_MAIN:
        main()

    if RUN_DRAW:
        # draw_price_charts()
        draw_price_charts_paged()
